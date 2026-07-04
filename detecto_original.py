#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Detecto - Log-Datei Scanner für kritische/personenbezogene Daten
#
# MIT License
#
# Copyright (c) 2026 Alexander Kornbrust
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import argparse
import configparser
import csv
import json
import re
import sys
import glob
import os
import unicodedata
from collections import OrderedDict

VERSION = "0.9"
INI_DATEI = "detecto.ini"

# Hartcodierte Fallback-Defaults (falls keine INI vorhanden)
MAX_BEISPIELE = 3
MIN_LAENGE = 5
MAX_CRITICAL = 5
REGEXP_DATEI = "regexp.csv"
FIELD_DATEI = "field.csv"
SUCHMUSTER_DATEI = "suchmuster.csv"
SUCHMUSTER_VERZEICHNIS = "suchmuster"
STOPWORD_REGEXP = "stopword_regexp.txt"
STOPWORD_FIELD = "stopword_field.txt"
STOPWORD_SUCHMUSTER = "stopword_suchmuster.txt"


def lade_ini(basisverzeichnis):
    """Lädt die detecto.ini und gibt ein Dict mit Default-Werten zurück."""
    ini_pfad = os.path.join(basisverzeichnis, INI_DATEI)
    config = configparser.ConfigParser()
    defaults = {
        "examplecount": MAX_BEISPIELE,
        "minlen": MIN_LAENGE,
        "critical": MAX_CRITICAL,
        "anon": False,
        "full": False,
        "nocolor": False,
        "showskipped": False,
        "anon_muster": ANON_MUSTER_DEFAULT,
        "search_regexp": True,
        "search_field": True,
        "search_suchmuster": True,
        "regexp": REGEXP_DATEI,
        "field": FIELD_DATEI,
        "suchmuster": SUCHMUSTER_DATEI,
        "suchmuster_verzeichnis": SUCHMUSTER_VERZEICHNIS,
        "stopword_regexp": STOPWORD_REGEXP,
        "stopword_field": STOPWORD_FIELD,
        "stopword_suchmuster": STOPWORD_SUCHMUSTER,
    }
    if os.path.isfile(ini_pfad):
        config.read(ini_pfad, encoding="utf-8")
        if config.has_section("defaults"):
            defaults["examplecount"] = config.getint("defaults", "examplecount",
                                                      fallback=MAX_BEISPIELE)
            defaults["minlen"] = config.getint("defaults", "minlen",
                                                fallback=MIN_LAENGE)
            defaults["critical"] = config.getint("defaults", "critical",
                                                  fallback=MAX_CRITICAL)
            defaults["anon"] = config.getboolean("defaults", "anon",
                                                  fallback=False)
            defaults["full"] = config.getboolean("defaults", "full",
                                                  fallback=False)
            defaults["nocolor"] = config.getboolean("defaults", "nocolor",
                                                     fallback=False)
            defaults["showskipped"] = config.getboolean("defaults", "showskipped",
                                                         fallback=False)
            defaults["anon_muster"] = config.get("defaults", "anon_muster",
                                                  fallback=ANON_MUSTER_DEFAULT)
            defaults["search_regexp"] = config.getboolean("defaults", "search_regexp",
                                                           fallback=True)
            defaults["search_field"] = config.getboolean("defaults", "search_field",
                                                          fallback=True)
            defaults["search_suchmuster"] = config.getboolean("defaults", "search_suchmuster",
                                                               fallback=True)
        if config.has_section("files"):
            defaults["regexp"] = config.get("files", "regexp",
                                             fallback=REGEXP_DATEI)
            defaults["field"] = config.get("files", "field",
                                            fallback=FIELD_DATEI)
            defaults["suchmuster"] = config.get("files", "suchmuster",
                                                 fallback=SUCHMUSTER_DATEI)
            defaults["suchmuster_verzeichnis"] = config.get(
                "files", "suchmuster_verzeichnis",
                fallback=SUCHMUSTER_VERZEICHNIS)
            defaults["stopword_regexp"] = config.get("files", "stopword_regexp",
                                                      fallback=STOPWORD_REGEXP)
            defaults["stopword_field"] = config.get("files", "stopword_field",
                                                     fallback=STOPWORD_FIELD)
            defaults["stopword_suchmuster"] = config.get(
                "files", "stopword_suchmuster",
                fallback=STOPWORD_SUCHMUSTER)
    return defaults


def normalisieren(text):
    """Normalisiert einen String: Kleinbuchstaben und Umlaute/Akzente entfernen.
    z.B. 'Müller' -> 'muller', 'č' -> 'c', 'ß' -> 'ss'
    """
    text = text.lower()
    text = text.replace("ß", "ss")
    # Unicode-Normalisierung: Zeichen in Basis + Akzent zerlegen, dann Akzente entfernen
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text


def lade_stopwords(dateipfad):
    """Lädt Stopwörter aus einer Textdatei (ein Wort pro Zeile).
    Gibt ein Set von normalisierten Stopwörtern zurück.
    """
    stopwords = set()
    if not os.path.isfile(dateipfad):
        return stopwords
    with open(dateipfad, "r", encoding="utf-8") as f:
        for zeile in f:
            wort = zeile.strip()
            if wort:
                stopwords.add(normalisieren(wort))
    return stopwords


KRIT_LABELS = {1: "kritisch", 2: "hoch", 3: "mittel", 4: "niedrig", 5: "info"}


def lade_regexp(dateipfad):
    """Lädt Regexp-Muster aus der CSV-Datei.
    Format: name::kritikalitaet::beschreibung::regexp
    Gibt eine Liste von Tupeln (name, krit, beschreibung, kompiliertes_pattern) zurück.
    """
    muster = []
    with open(dateipfad, "r", encoding="utf-8") as f:
        for zeile in f:
            zeile = zeile.strip()
            if not zeile:
                continue
            teile = zeile.split("::")
            if len(teile) < 4:
                print(f"WARNUNG: Ungültiges Format in {dateipfad}: {zeile}")
                continue
            name, krit, beschreibung, pattern = teile[0], int(teile[1]), teile[2], teile[3]
            try:
                kompiliert = re.compile(pattern)
                muster.append((name, krit, beschreibung, kompiliert))
            except re.error as e:
                print(f"WARNUNG: Ungültiger Regexp '{pattern}': {e}")
    return muster


def lade_field_muster(dateipfad):
    """Lädt Field-Muster aus der CSV-Datei.
    Format: name::kritikalitaet::beschreibung::regexp[::offset]
    Offset gibt an, wie viele Token nach dem Match der Wert steht (Standard: 1).
    Gibt eine Liste von Tupeln (name, krit, beschreibung, kompiliertes_pattern, offset) zurück.
    """
    muster = []
    with open(dateipfad, "r", encoding="utf-8") as f:
        for zeile in f:
            zeile = zeile.strip()
            if not zeile:
                continue
            teile = zeile.split("::")
            if len(teile) < 4:
                print(f"WARNUNG: Ungültiges Format in {dateipfad}: {zeile}")
                continue
            name, krit, beschreibung, pattern = teile[0], int(teile[1]), teile[2], teile[3]
            offset = int(teile[4]) if len(teile) >= 5 else 1
            try:
                kompiliert = re.compile(pattern)
                muster.append((name, krit, beschreibung, kompiliert, offset))
            except re.error as e:
                print(f"WARNUNG: Ungültiger Regexp '{pattern}': {e}")
    return muster


def lade_suchmuster(dateipfad, verzeichnis, minlen=MIN_LAENGE):
    """Lädt Suchmuster aus der Indexdatei und den referenzierten CSV-Dateien.
    Format der Indexdatei: Name:dateiname.csv
    Gibt eine Liste von Tupeln (name, set_von_werten) zurück.
    """
    muster = []
    with open(dateipfad, "r", encoding="utf-8") as f:
        for zeile in f:
            zeile = zeile.strip()
            if not zeile:
                continue
            teile = zeile.split("::")
            if len(teile) < 3:
                print(f"WARNUNG: Ungültiges Format in {dateipfad}: {zeile}")
                continue
            name, krit, dateiname = teile[0], int(teile[1]), teile[2]
            csv_pfad = os.path.join(verzeichnis, dateiname)
            if not os.path.isfile(csv_pfad):
                print(f"WARNUNG: Datei nicht gefunden: {csv_pfad}")
                continue
            werte = set()
            with open(csv_pfad, "r", encoding="utf-8") as csv_f:
                reader = csv.reader(csv_f)
                header = next(reader, None)  # Header überspringen
                for row in reader:
                    if not row:
                        continue
                    # Erste Spalte enthält den Suchwert
                    wert = row[0].strip()
                    if wert and len(wert) >= minlen:
                        werte.add(normalisieren(wert))
            if werte:
                muster.append((name, krit, werte))
    return muster


def json_werte_extrahieren(obj):
    """Extrahiert rekursiv alle String-Werte aus einem JSON-Objekt."""
    werte = []
    if isinstance(obj, dict):
        for v in obj.values():
            werte.extend(json_werte_extrahieren(v))
    elif isinstance(obj, list):
        for v in obj:
            werte.extend(json_werte_extrahieren(v))
    elif isinstance(obj, str):
        werte.append(obj)
    return werte


FIELD_SEPARATOREN = {"->", "=>", ":", "=", "==", "-->", "<<", ">>"}


def finde_field_wert(token_liste, start_pos):
    """Findet den eigentlichen Wert nach einem Field-Match.
    Ueberspringt Separatoren (->  =>  :  =) und leere Wrapper-Tokens.
    Schaut maximal 3 Tokens voraus.
    Gibt (wert, original_token) zurueck oder (None, None).
    """
    max_lookahead = min(start_pos + 4, len(token_liste))
    for j in range(start_pos, max_lookahead):
        kandidat = token_liste[j]
        # Separatoren ueberspringen
        if kandidat in FIELD_SEPARATOREN:
            continue
        # Reine Interpunktion/Wrapper ueberspringen
        bereinigt = kandidat.strip('"\'{}[](),.:;!?')
        if not bereinigt:
            continue
        return bereinigt, kandidat
    return None, None


def extrahiere_json_fragmente(text):
    """Extrahiert eingebettete JSON-Objekte {...} aus einem Text.
    Verwendet Klammer-Balancing um vollstaendige Objekte zu finden.
    Gibt eine Liste von (json_string, start, end) Tupeln zurueck.
    """
    fragmente = []
    i = 0
    while i < len(text):
        if text[i] == '{':
            tiefe = 0
            start = i
            in_string = False
            escape = False
            for j in range(i, len(text)):
                ch = text[j]
                if escape:
                    escape = False
                    continue
                if ch == '\\' and in_string:
                    escape = True
                    continue
                if ch == '"' and not escape:
                    in_string = not in_string
                    continue
                if in_string:
                    continue
                if ch == '{':
                    tiefe += 1
                elif ch == '}':
                    tiefe -= 1
                    if tiefe == 0:
                        fragmente.append((text[start:j + 1], start, j + 1))
                        i = j + 1
                        break
            else:
                # Unbalancierte Klammer - Rest ueberspringen
                i += 1
        else:
            i += 1
    return fragmente


def tokenize(zeile):
    """Splittet eine Log-Zeile in einzelne Token auf.
    Unterstützt Plain-Text-Logs und JSON-Logs (eine JSON-Zeile pro Logzeile).
    Trennt an Whitespace, URL-Trennzeichen (?, &, =) und weiteren Sonderzeichen.
    Eingebettete JSON-Fragmente werden als Ganzes erhalten und zusaetzlich
    intern tokenisiert.
    """
    # JSON-Log erkennen: Zeile beginnt mit '{'
    text_teile = [zeile]
    if zeile.startswith("{"):
        try:
            json_obj = json.loads(zeile)
            text_teile = json_werte_extrahieren(json_obj)
        except (json.JSONDecodeError, ValueError):
            pass  # Kein gültiges JSON, als Plain-Text behandeln

    ergebnis = []
    for text in text_teile:
        # Eingebettete JSON-Fragmente extrahieren und als eigene Tokens erhalten
        fragmente = extrahiere_json_fragmente(text)
        if fragmente:
            # Nicht-JSON-Teile und JSON-Fragmente getrennt verarbeiten
            letztes_ende = 0
            for frag_text, frag_start, frag_end in fragmente:
                # Text vor dem Fragment normal tokenisieren
                vor_fragment = text[letztes_ende:frag_start]
                if vor_fragment.strip():
                    token = re.split(r'[\s?&=,;|]+', vor_fragment)
                    ergebnis.extend(t for t in token if t)
                # JSON-Fragment als Ganzes hinzufuegen
                ergebnis.append(frag_text)
                # Zusaetzlich den Inhalt des Fragments tokenisieren
                try:
                    json_obj = json.loads(frag_text)
                    for wert in json_werte_extrahieren(json_obj):
                        inner_tokens = re.split(r'[\s?&=,;|]+', wert)
                        ergebnis.extend(t for t in inner_tokens if t)
                except (json.JSONDecodeError, ValueError):
                    # Kein gueltiges JSON, intern normal splitten
                    inner_tokens = re.split(r'[\s?&=,;|]+', frag_text)
                    ergebnis.extend(t for t in inner_tokens if t)
                letztes_ende = frag_end
            # Rest nach letztem Fragment
            rest = text[letztes_ende:]
            if rest.strip():
                token = re.split(r'[\s?&=,;|]+', rest)
                ergebnis.extend(t for t in token if t)
        else:
            # Kein JSON-Fragment, normal tokenisieren
            token = re.split(r'[\s?&=,;|]+', text)
            ergebnis.extend(t for t in token if t)

        # URL-Parameterwerte als Ganzes extrahieren (key=value bis zum nächsten &)
        # Damit werden auch mehrteilige Werte wie "65 170839 J 08 8" erfasst
        for param_wert in re.findall(r'[?&](\w+=([^&\n]+))', text):
            wert = param_wert[1].strip()
            if wert and wert not in ergebnis:
                ergebnis.append(wert)

    return ergebnis


def sortiere_ergebnisse(ergebnisse):
    """Sortiert Ergebnisse stabil nach Kritikalität, Typ und Name."""
    art_reihenfolge = {"field": 0, "regexp": 1, "string": 2}
    return sorted(
        ergebnisse.items(),
        key=lambda eintrag: (
            eintrag[1][1],
            art_reihenfolge.get(eintrag[1][0], 9),
            eintrag[0],
        ),
    )


def analysiere_logdateien(logdateien, regexp_muster, field_muster, such_muster,
                          sw_regexp=None, sw_field=None, sw_suchmuster=None):
    """Analysiert die Logdateien und sucht nach Matches.
    Gibt ein Tupel (ergebnisse, anzahl_zeilen) zurück.
    ergebnisse ist ein OrderedDict mit name -> (typ, dict(wert -> [zeilen])).
    """
    ergebnisse = OrderedDict()
    anzahl_zeilen = 0
    sw_regexp = sw_regexp or set()
    sw_field = sw_field or set()
    sw_suchmuster = sw_suchmuster or set()

    for name, krit, _, _ in regexp_muster:
        ergebnisse[name] = ("regexp", krit, {})
    for name, krit, _, _, _ in field_muster:
        ergebnisse[name] = ("field", krit, {})
    for name, krit, _ in such_muster:
        ergebnisse[name] = ("string", krit, {})

    for logdatei in logdateien:
        dateiname = os.path.basename(logdatei)
        try:
            with open(logdatei, "r", encoding="utf-8", errors="replace") as f:
                for zeile in f:
                    anzahl_zeilen += 1
                    zeile = zeile.strip()
                    if not zeile:
                        continue
                    token_liste = tokenize(zeile)
                    for i, token in enumerate(token_liste):
                        for name, _, _, pattern in regexp_muster:
                            if pattern.search(token):
                                if normalisieren(token) in sw_regexp:
                                    continue
                                treffer = ergebnisse[name][2]
                                if token not in treffer:
                                    treffer[token] = []
                                treffer[token].append((dateiname, zeile, None))
                        # Field-Muster: Token matched Feldname,
                        # der eigentliche Wert folgt nach Separatoren
                        for name, _, _, pattern, offset in field_muster:
                            if pattern.search(token):
                                wert, _ = finde_field_wert(
                                    token_liste, i + offset)
                                if not wert:
                                    continue
                                # Pfade und Sternchen-Folgen ignorieren
                                if wert.startswith(("/", "\\", "~/")) or \
                                   (len(wert) >= 2 and wert[1] == ":") or \
                                   re.fullmatch(r'\*+', wert):
                                    continue
                                if normalisieren(wert) in sw_field:
                                    continue
                                treffer = ergebnisse[name][2]
                                if wert not in treffer:
                                    treffer[wert] = []
                                treffer[wert].append((dateiname, zeile, token))
                        for name, _, werte in such_muster:
                            if normalisieren(token) in werte:
                                if normalisieren(token) in sw_suchmuster:
                                    continue
                                treffer = ergebnisse[name][2]
                                if token not in treffer:
                                    treffer[token] = []
                                treffer[token].append((dateiname, zeile, None))
        except (IOError, OSError) as e:
            print(f"WARNUNG: Fehler beim Lesen von {logdatei}: {e}")

    return ergebnisse, anzahl_zeilen


ANON_MUSTER_DEFAULT = "sss**sss**sss**"
ANON_GRUPPEN = None  # Wird in main() aus INI gesetzt


def parse_anon_muster(muster):
    """Parst das Anonymisierungsmuster in eine Liste von (typ, laenge) Tupeln.
    typ='s' = Originalzeichen, typ='*' = Stern.
    Gibt None zurueck wenn das Muster '<redacted>' ist.
    """
    if muster.strip().lower() == "<redacted>":
        return None
    gruppen = []
    i = 0
    while i < len(muster):
        ch = muster[i]
        if ch == 's':
            count = 0
            while i < len(muster) and muster[i] == 's':
                count += 1
                i += 1
            gruppen.append(('s', count))
        elif ch in ('*', 'x'):
            count = 0
            while i < len(muster) and muster[i] in ('*', 'x'):
                count += 1
                i += 1
            gruppen.append(('*', count))
        else:
            i += 1
    # Leeres Ergebnis (ungueltiges Muster) -> Default verwenden
    if not gruppen:
        return parse_anon_muster(ANON_MUSTER_DEFAULT)
    return gruppen


def redact(text, muster_gruppen=None):
    """Anonymisiert einen String nach dem konfigurierten Muster.
    muster_gruppen: Liste von (typ, laenge) Tupeln aus parse_anon_muster(),
                    oder None fuer '<redacted>'.
    """
    if muster_gruppen is None:
        return "<redacted>"
    if not muster_gruppen:
        return "<redacted>"
    ergebnis = []
    i = 0
    g = 0
    while i < len(text):
        typ, laenge = muster_gruppen[g % len(muster_gruppen)]
        if typ == 's':
            ergebnis.append(text[i:i + laenge])
            i += laenge
        else:
            verbleibend = min(laenge, len(text) - i)
            ergebnis.append("*" * verbleibend)
            i += verbleibend
        g += 1
    return "".join(ergebnis)[:len(text)]


FARBE_ROT = "\033[91m"
FARBE_GRUEN = "\033[92m"
FARBE_DUNKELGRUEN = "\033[32m"
FARBE_GELB = "\033[93m"
FARBE_RESET = "\033[0m"

FARBE_TYP = {
    "regexp": FARBE_GRUEN,
    "field": FARBE_GELB,
    "string": FARBE_GRUEN,
}


def hervorheben(zeile, token, anon=False, farbe=True, feld_token=None):
    """Hebt alle Vorkommen von token in der Zeile hervor (case-insensitive).
    Bei anon=True wird der Token zusätzlich redacted dargestellt.
    Bei farbe=False werden keine ANSI-Farbcodes verwendet (für Logdatei).
    Bei feld_token wird zusätzlich der Feldname in Dunkelgrün hervorgehoben.
    """
    anzeige = redact(token, ANON_GRUPPEN) if anon else token
    pattern = re.compile(re.escape(token), re.IGNORECASE)
    if farbe:
        ergebnis = pattern.sub(f"{FARBE_ROT}{anzeige}{FARBE_RESET}", zeile)
        if feld_token:
            feld_pattern = re.compile(re.escape(feld_token), re.IGNORECASE)
            ergebnis = feld_pattern.sub(
                f"{FARBE_DUNKELGRUEN}{feld_token}{FARBE_RESET}", ergebnis)
        return ergebnis
    else:
        return pattern.sub(anzeige, zeile)


def erzeuge_ergebnis_zeilen(ergebnisse, beispiel_anzahl=MAX_BEISPIELE,
                            anon=False, full=False, farbe=True, critical=5,
                            showskipped=False):
    """Erzeugt die Ergebniszeilen als Liste von Strings.
    farbe=False unterdrückt ANSI-Farbcodes (für Logdatei-Ausgabe).
    critical: Nur Findings mit Kritikalität <= critical anzeigen.
    showskipped: Auch Muster ohne Treffer anzeigen.
    """
    zeilen = []
    zeilen.append("")
    zeilen.append("=== Detecto - Ergebnisse ===")
    zeilen.append("")
    gefunden = False

    def formatierte_kopfzeile(typ, krit, name, inhalt):
        krit_label = KRIT_LABELS.get(krit, str(krit))
        if farbe:
            typ_farbe = FARBE_TYP.get(typ, "")
            krit_farbe = FARBE_ROT if krit <= 2 else FARBE_GELB if krit == 3 else FARBE_RESET
            return (
                f"{typ_farbe}[{typ}]{FARBE_RESET} "
                f"{krit_farbe}[{krit_label}]{FARBE_RESET} "
                f"{FARBE_GELB}{name}{FARBE_RESET}: {inhalt}"
            )
        return f"[{typ}] [{krit_label}] {name}: {inhalt}"

    for name, (typ, krit, treffer) in ergebnisse.items():
        if krit > critical:
            continue
        if treffer:
            gefunden = True
            beispiel_keys = sorted(treffer.keys())[:beispiel_anzahl]
            if anon:
                anzeige = [redact(b, ANON_GRUPPEN) for b in beispiel_keys]
            else:
                anzeige = beispiel_keys
            zeilen.append(formatierte_kopfzeile(typ, krit, name, ", ".join(anzeige)))
            if full:
                for token in beispiel_keys:
                    eintrag = treffer[token][0]
                    dateiname = eintrag[0]
                    erste_zeile = eintrag[1]
                    feld_token = eintrag[2] if len(eintrag) > 2 else None
                    markiert = hervorheben(erste_zeile, token, anon, farbe,
                                           feld_token)
                    zeilen.append(f"  → [{dateiname}] {markiert}")
                zeilen.append("")
        elif showskipped:
            zeilen.append(formatierte_kopfzeile(typ, krit, name, "<nothing found>"))
    if not gefunden and not showskipped:
        zeilen.append("Keine kritischen Daten gefunden.")
    zeilen.append("")
    return zeilen


def ausgabe(ergebnisse, beispiel_anzahl=MAX_BEISPIELE, anon=False, full=False,
            nocolor=False, critical=5, showskipped=False):
    """Gibt die Zusammenfassung der Findings auf stdout aus."""
    zeilen = erzeuge_ergebnis_zeilen(ergebnisse, beispiel_anzahl, anon, full,
                                     farbe=not nocolor, critical=critical,
                                     showskipped=showskipped)
    for zeile in zeilen:
        print(zeile)


def exportiere_xlsx(dateiname, ergebnisse, beispiel_anzahl, critical,
                    aufruf, logdateien, anzahl_zeilen, dauer_text,
                    regexp_muster, field_muster, such_muster, full=False,
                    excelanon=False):
    """Exportiert die Findings als Excel-Datei fuer die Kundenbearbeitung."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("FEHLER: openpyxl ist nicht installiert. "
              "Bitte installieren mit: pip3 install openpyxl")
        return

    from datetime import datetime

    wb = Workbook()

    # === Sheet 1: Findings ===
    ws_findings = wb.active
    ws_findings.title = "Findings"

    # Header-Formatierung
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                              fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    def schreibe_header(worksheet, headers, row=1):
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def setze_zeilenrahmen(worksheet, row, von, bis):
        for col in range(von, bis + 1):
            worksheet.cell(row=row, column=col).border = thin_border

    def markiere_logeintrag(log_zeile, wert, feld_token=None):
        markierter_log = log_zeile
        if feld_token:
            feld_pat = re.compile(re.escape(feld_token), re.IGNORECASE)
            markierter_log = feld_pat.sub(f"[FELD:{feld_token}]", markierter_log)
        wert_pat = re.compile(re.escape(wert), re.IGNORECASE)
        ersatz = redact(wert, ANON_GRUPPEN) if excelanon else wert
        return wert_pat.sub(f">>>{ersatz}<<<", markierter_log)

    # Header schreiben
    headers = ["Bearbeiter", "Finding", "Kommentar",
               "Art", "Muster", "Kritikalität (1-5)",
               "Beispiele", "gefunden in"]
    schreibe_header(ws_findings, headers)

    # Kundenbearbeitungs-Spalten hellgelb hinterlegen
    kunden_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                              fill_type="solid")

    # Kritikalitaets-Farben
    krit_fills = {
        1: PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
        2: PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),
        3: PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
        4: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        5: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    }

    # Daten sortiert schreiben: Kritikalitaet (1..5), Art (field, regexp, string), Name
    sortierte_ergebnisse = sortiere_ergebnisse(ergebnisse)

    zeile_nr = 2
    for name, (typ, krit, treffer) in sortierte_ergebnisse:
        if not treffer or krit > critical:
            continue

        beispiel_keys = sorted(treffer.keys())[:beispiel_anzahl]
        if excelanon:
            beispiele_str = ", ".join(redact(b, ANON_GRUPPEN) for b in beispiel_keys)
        else:
            beispiele_str = ", ".join(beispiel_keys)

        # Logdateien sammeln, in denen Treffer gefunden wurden
        log_set = set()
        for wert_treffer in treffer.values():
            for eintrag in wert_treffer:
                dateiname_log = eintrag[0]
                log_set.add(dateiname_log)
        logdateien_str = ", ".join(sorted(log_set))

        # Zeile schreiben
        ws_findings.cell(row=zeile_nr, column=1, value="").fill = kunden_fill
        ws_findings.cell(row=zeile_nr, column=2, value="").fill = kunden_fill
        ws_findings.cell(row=zeile_nr, column=3, value="").fill = kunden_fill
        ws_findings.cell(row=zeile_nr, column=4, value=typ)
        ws_findings.cell(row=zeile_nr, column=5, value=name)
        krit_cell = ws_findings.cell(row=zeile_nr, column=6, value=krit)
        krit_cell.fill = krit_fills.get(krit, krit_fills[5])
        krit_cell.alignment = Alignment(horizontal="center")
        ws_findings.cell(row=zeile_nr, column=7,
                         value=beispiele_str).alignment = Alignment(
                             wrap_text=True, vertical="top")
        ws_findings.cell(row=zeile_nr, column=8,
                         value=logdateien_str).alignment = Alignment(
                             wrap_text=True, vertical="top")

        # Rahmen fuer alle Zellen der Zeile
        setze_zeilenrahmen(ws_findings, zeile_nr, 1, 8)

        zeile_nr += 1

    # Spaltenbreiten
    ws_findings.column_dimensions["A"].width = 15
    ws_findings.column_dimensions["B"].width = 15
    ws_findings.column_dimensions["C"].width = 25
    ws_findings.column_dimensions["D"].width = 10
    ws_findings.column_dimensions["E"].width = 20
    ws_findings.column_dimensions["F"].width = 18
    ws_findings.column_dimensions["G"].width = 60
    ws_findings.column_dimensions["H"].width = 53

    # Autofilter
    ws_findings.auto_filter.ref = f"A1:H{max(zeile_nr - 1, 1)}"

    # Erste Zeile fixieren
    ws_findings.freeze_panes = "A2"

    # === Sheet 2: Tool ===
    ws_tool = wb.create_sheet(title="Tool")

    tool_header_font = Font(bold=True, size=11)
    tool_daten = [
        ("Detecto Version", f"v{VERSION}"),
        ("Datum", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Aufruf", "python3 detecto.py " + aufruf),
        ("", ""),
        ("Analysierte Logdateien", str(len(logdateien))),
        ("Logdateien", ", ".join(logdateien)),
        ("Anzahl der Zeilen", str(anzahl_zeilen)),
        ("Analyse-Dauer", dauer_text),
        ("", ""),
        ("Geladene Regexp-Muster", str(len(regexp_muster))),
        ("Geladene Field-Muster", str(len(field_muster))),
        ("Geladene Suchmuster-Kategorien", str(len(such_muster))),
        ("Critical-Filter", str(critical)),
        ("Beispiele pro Typ", str(beispiel_anzahl)),
    ]
    for row_nr, (label, wert) in enumerate(tool_daten, 1):
        label_cell = ws_tool.cell(row=row_nr, column=1, value=label)
        label_cell.font = tool_header_font
        ws_tool.cell(row=row_nr, column=2, value=wert)

    ws_tool.column_dimensions["A"].width = 30
    ws_tool.column_dimensions["B"].width = 60

    # === Sheet 3: Regexp ===
    ws_regexp = wb.create_sheet(title="Regexp")
    regexp_headers = ["Name", "Kritikalität", "Beschreibung", "Pattern"]
    schreibe_header(ws_regexp, regexp_headers)
    for row_nr, (name, krit, beschreibung, pattern) in enumerate(regexp_muster, 2):
        ws_regexp.cell(row=row_nr, column=1, value=name).border = thin_border
        krit_cell = ws_regexp.cell(row=row_nr, column=2, value=krit)
        krit_cell.fill = krit_fills.get(krit, krit_fills[5])
        krit_cell.alignment = Alignment(horizontal="center")
        krit_cell.border = thin_border
        ws_regexp.cell(row=row_nr, column=3, value=beschreibung).border = thin_border
        ws_regexp.cell(row=row_nr, column=4, value=pattern.pattern).border = thin_border
    ws_regexp.column_dimensions["A"].width = 20
    ws_regexp.column_dimensions["B"].width = 14
    ws_regexp.column_dimensions["C"].width = 40
    ws_regexp.column_dimensions["D"].width = 60
    ws_regexp.freeze_panes = "A2"
    ws_regexp.auto_filter.ref = f"A1:D{max(len(regexp_muster) + 1, 1)}"

    # === Sheet 4: Field ===
    ws_field = wb.create_sheet(title="Field")
    field_headers = ["Name", "Kritikalität", "Beschreibung", "Pattern", "Offset"]
    schreibe_header(ws_field, field_headers)
    for row_nr, (name, krit, beschreibung, pattern, offset) in enumerate(field_muster, 2):
        ws_field.cell(row=row_nr, column=1, value=name).border = thin_border
        krit_cell = ws_field.cell(row=row_nr, column=2, value=krit)
        krit_cell.fill = krit_fills.get(krit, krit_fills[5])
        krit_cell.alignment = Alignment(horizontal="center")
        krit_cell.border = thin_border
        ws_field.cell(row=row_nr, column=3, value=beschreibung).border = thin_border
        ws_field.cell(row=row_nr, column=4, value=pattern.pattern).border = thin_border
        ws_field.cell(row=row_nr, column=5, value=offset).border = thin_border
        ws_field.cell(row=row_nr, column=5).alignment = Alignment(horizontal="center")
    ws_field.column_dimensions["A"].width = 20
    ws_field.column_dimensions["B"].width = 14
    ws_field.column_dimensions["C"].width = 35
    ws_field.column_dimensions["D"].width = 60
    ws_field.column_dimensions["E"].width = 10
    ws_field.freeze_panes = "A2"
    ws_field.auto_filter.ref = f"A1:E{max(len(field_muster) + 1, 1)}"

    # === Sheet 5: String ===
    ws_string = wb.create_sheet(title="String")
    string_headers = ["Name", "Kritikalität", "Datei", "Einträge"]
    schreibe_header(ws_string, string_headers)
    for row_nr, (name, krit, werte) in enumerate(such_muster, 2):
        ws_string.cell(row=row_nr, column=1, value=name).border = thin_border
        krit_cell = ws_string.cell(row=row_nr, column=2, value=krit)
        krit_cell.fill = krit_fills.get(krit, krit_fills[5])
        krit_cell.alignment = Alignment(horizontal="center")
        krit_cell.border = thin_border
        ws_string.cell(row=row_nr, column=3, value=name.lower() + ".csv"
                        if not name.endswith(".csv") else name).border = thin_border
        ws_string.cell(row=row_nr, column=4, value=len(werte)).border = thin_border
        ws_string.cell(row=row_nr, column=4).alignment = Alignment(horizontal="center")
    ws_string.column_dimensions["A"].width = 20
    ws_string.column_dimensions["B"].width = 14
    ws_string.column_dimensions["C"].width = 25
    ws_string.column_dimensions["D"].width = 12
    ws_string.freeze_panes = "A2"
    ws_string.auto_filter.ref = f"A1:D{max(len(such_muster) + 1, 1)}"

    # === Sheet "Full" direkt nach Findings (nur wenn --full aktiv) ===
    if full:
        ws_full = wb.create_sheet(title="Full", index=1)
        full_headers = ["Logdatei", "Art", "Muster", "Kritikalität (1-5)",
                        "Gefundener Wert", "Feld", "Logeintrag"]
        schreibe_header(ws_full, full_headers)

        # Sortiert wie Findings-Sheet
        sortierte_full = sortiere_ergebnisse(ergebnisse)

        wert_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC",
                                fill_type="solid")
        feld_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE",
                                fill_type="solid")

        full_zeile = 2
        for name, (typ, krit, treffer) in sortierte_full:
            if not treffer or krit > critical:
                continue
            beispiel_keys = sorted(treffer.keys())[:beispiel_anzahl]
            for wert in beispiel_keys:
                eintrag = treffer[wert][0]
                dateiname_log = eintrag[0]
                log_zeile = eintrag[1]
                feld_token = eintrag[2] if len(eintrag) > 2 else None
                anzeige_wert = redact(wert, ANON_GRUPPEN) if excelanon else wert

                ws_full.cell(row=full_zeile, column=1,
                             value=dateiname_log).border = thin_border
                ws_full.cell(row=full_zeile, column=2,
                             value=typ).border = thin_border
                ws_full.cell(row=full_zeile, column=3,
                             value=name).border = thin_border
                krit_cell = ws_full.cell(row=full_zeile, column=4, value=krit)
                krit_cell.fill = krit_fills.get(krit, krit_fills[5])
                krit_cell.alignment = Alignment(horizontal="center")
                krit_cell.border = thin_border

                # Gefundener Wert mit rotem Hintergrund
                wert_cell = ws_full.cell(row=full_zeile, column=5,
                                         value=anzeige_wert)
                wert_cell.fill = wert_fill
                wert_cell.font = Font(bold=True)
                wert_cell.border = thin_border

                # Feld-Token mit gruenem Hintergrund
                feld_cell = ws_full.cell(row=full_zeile, column=6,
                                         value=feld_token or "")
                if feld_token:
                    feld_cell.fill = feld_fill
                    feld_cell.font = Font(bold=True)
                feld_cell.border = thin_border

                # Logeintrag als Plain-Text mit Markierungen
                try:
                    ws_full.cell(row=full_zeile, column=7,
                                 value=markiere_logeintrag(log_zeile, wert, feld_token))
                except Exception:
                    ws_full.cell(row=full_zeile, column=7,
                                 value=log_zeile)

                ws_full.cell(row=full_zeile, column=7).border = thin_border
                ws_full.cell(row=full_zeile, column=7).alignment = Alignment(
                    wrap_text=True, vertical="top")
                full_zeile += 1

        ws_full.column_dimensions["A"].width = 25
        ws_full.column_dimensions["B"].width = 10
        ws_full.column_dimensions["C"].width = 20
        ws_full.column_dimensions["D"].width = 18
        ws_full.column_dimensions["E"].width = 30
        ws_full.column_dimensions["F"].width = 25
        ws_full.column_dimensions["G"].width = 100
        ws_full.freeze_panes = "A2"
        ws_full.auto_filter.ref = f"A1:G{max(full_zeile - 1, 1)}"

    # Speichern
    wb.save(dateiname)
    anon_hinweis = " (anonymisiert)" if excelanon else ""
    print(f"Excel-Report gespeichert in: {dateiname}{anon_hinweis}")


def finde_logdateien(muster):
    """Findet Logdateien basierend auf dem Muster (z.B. test.log oder *.log)."""
    dateien = glob.glob(muster)
    return [d for d in dateien if os.path.isfile(d)]


def parse_args(ini_defaults):
    """Parst die Kommandozeilenargumente mit Defaults aus der INI-Datei."""
    parser = argparse.ArgumentParser(
        prog="detecto",
        description="Detecto - Scannt Log-Dateien (Websphere/Liberty) nach "
                    "kritischen und personenbezogenen Daten.",
        epilog="Beispiele:\n"
               "  python3 detecto.py test.log\n"
               "  python3 detecto.py \"*.log\"\n"
               "  python3 detecto.py test.log --examplecount=5\n"
               "\n"
               "Dokumentation:\n"
               "  Readme:          readme.md\n"
               "  Changelog:       changelog.md\n"
               "  Erste Schritte:  firststeps.md\n",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "logdateien",
        nargs="*",
        help="Log-Datei(en) zum Scannen, z.B. test.log, *.log oder "
             "log*.log (mehrere Dateien möglich)",
    )
    parser.add_argument(
        "--examplecount",
        type=int,
        default=ini_defaults["examplecount"],
        metavar="N",
        help="Anzahl der unterschiedlichen Beispiele pro Fundtyp "
             f"(Standard: {ini_defaults['examplecount']})",
    )
    parser.add_argument(
        "--anon",
        action="store_true",
        default=ini_defaults["anon"],
        help="Beispiele redacted ausgeben (sss**sss**sss**...)",
    )
    parser.add_argument(
        "--full",
        action="store_true",
        default=ini_defaults["full"],
        help="Zeigt zusätzlich die Log-Zeilen der Findings mit "
             "farblicher Hervorhebung an",
    )
    parser.add_argument(
        "--minlen",
        type=int,
        default=ini_defaults["minlen"],
        metavar="N",
        help="Mindestlänge für Suchmuster-Strings, um False Positives "
             f"zu vermeiden (Standard: {ini_defaults['minlen']}, Minimum: 2)",
    )
    parser.add_argument(
        "--status",
        action="store_true",
        default=False,
        help="Zeigt eine Zusammenfassung der geladenen Suchdaten an",
    )
    parser.add_argument(
        "--nocolor",
        action="store_true",
        default=ini_defaults["nocolor"],
        help="Deaktiviert die farbliche Ausgabe komplett",
    )
    parser.add_argument(
        "--critical",
        type=int,
        default=ini_defaults["critical"],
        metavar="N",
        help="Zeigt nur Findings bis Kritikalitaet N an "
             "(1=kritisch, 2=hoch, 3=mittel, 4=niedrig, 5=alle). "
             f"Standard: {ini_defaults['critical']}",
    )
    parser.add_argument(
        "--showskipped",
        action="store_true",
        default=ini_defaults["showskipped"],
        help="Zeigt auch Muster ohne Treffer an (<nothing found>)",
    )
    parser.add_argument(
        "--logresult",
        nargs="?",
        const="",
        default=None,
        metavar="DATEI",
        help="Speichert das Ergebnis (nicht anonymisiert) in eine Datei. "
             "Ohne Dateiname: detecto_<timestamp>.log",
    )
    parser.add_argument(
        "--logresultanon",
        nargs="?",
        const="",
        default=None,
        metavar="DATEI",
        help="Speichert das Ergebnis anonymisiert in eine Datei. "
             "Ohne Dateiname: detecto_<timestamp>.log",
    )
    parser.add_argument(
        "--xlsx",
        nargs="?",
        const="",
        default=None,
        metavar="DATEI",
        help="Exportiert die Findings als Excel-Datei (fuer Kundenbearbeitung). "
             "Ohne Dateiname: detecto_<timestamp>.xlsx",
    )
    parser.add_argument(
        "--excelanon",
        action="store_true",
        default=False,
        help="Anonymisiert die kritischen Werte in der Excel-Datei "
             "(unabhaengig von --anon)",
    )
    args = parser.parse_args()
    if not args.logdateien and not args.status:
        parser.print_help()
        sys.exit(0)
    return args


def erzeuge_aufruf_string(args, logdateien, ini_defaults=None):
    """Erzeugt einen kompakten Aufruf-String mit den tatsaechlichen Parametern."""
    ini = ini_defaults or {}
    teile = []
    # Logdateien
    teile.extend(logdateien)
    # Nur Parameter anfuegen, die vom INI-Default abweichen oder gesetzt sind
    if args.examplecount != ini.get("examplecount", MAX_BEISPIELE):
        teile.append(f"--examplecount={args.examplecount}")
    if args.anon:
        teile.append("--anon")
    if args.full:
        teile.append("--full")
    if args.minlen != ini.get("minlen", MIN_LAENGE):
        teile.append(f"--minlen={args.minlen}")
    if args.critical != ini.get("critical", MAX_CRITICAL):
        teile.append(f"--critical={args.critical}")
    if args.nocolor:
        teile.append("--nocolor")
    if args.showskipped:
        teile.append("--showskipped")
    if args.logresult is not None:
        teile.append(f"--logresult={args.logresult}" if args.logresult else "--logresult")
    if args.logresultanon is not None:
        teile.append(f"--logresultanon={args.logresultanon}" if args.logresultanon else "--logresultanon")
    if args.xlsx is not None:
        teile.append(f"--xlsx={args.xlsx}" if args.xlsx else "--xlsx")
    if args.excelanon:
        teile.append("--excelanon")
    return " ".join(teile)


def print_header():
    """Gibt den Detecto-Header aus."""
    from datetime import date
    print(f"Detecto v{VERSION}  ({date.today().isoformat()})")
    print("Copyright (c) 2026 Alexander Kornbrust  |  MIT License")
    print()


def main():
    print_header()
    basisverzeichnis = os.path.dirname(os.path.abspath(__file__))
    ini_defaults = lade_ini(basisverzeichnis)
    args = parse_args(ini_defaults)

    log_eingaben = args.logdateien
    beispiel_anzahl = args.examplecount
    minlen = max(2, args.minlen)

    # Anonymisierungsmuster parsen und global setzen
    global ANON_GRUPPEN
    ANON_GRUPPEN = parse_anon_muster(ini_defaults["anon_muster"])

    regexp_pfad = os.path.join(basisverzeichnis, ini_defaults["regexp"])
    field_pfad = os.path.join(basisverzeichnis, ini_defaults["field"])
    suchmuster_pfad = os.path.join(basisverzeichnis, ini_defaults["suchmuster"])
    suchmuster_verz = os.path.join(basisverzeichnis,
                                    ini_defaults["suchmuster_verzeichnis"])

    if not os.path.isfile(regexp_pfad):
        print(f"FEHLER: Regexp-Datei nicht gefunden: {regexp_pfad}")
        sys.exit(1)
    if not os.path.isfile(suchmuster_pfad):
        print(f"FEHLER: Suchmuster-Datei nicht gefunden: {suchmuster_pfad}")
        sys.exit(1)

    regexp_muster = lade_regexp(regexp_pfad) if ini_defaults["search_regexp"] else []
    field_muster = lade_field_muster(field_pfad) if ini_defaults["search_field"] and os.path.isfile(field_pfad) else []
    such_muster = lade_suchmuster(suchmuster_pfad, suchmuster_verz, minlen) if ini_defaults["search_suchmuster"] else []

    sw_regexp = lade_stopwords(os.path.join(basisverzeichnis,
                                             ini_defaults["stopword_regexp"]))
    sw_field = lade_stopwords(os.path.join(basisverzeichnis,
                                            ini_defaults["stopword_field"]))
    sw_suchmuster = lade_stopwords(os.path.join(basisverzeichnis,
                                                 ini_defaults["stopword_suchmuster"]))

    if args.status:
        print(f"=== Detecto - Status ===\n")
        label_b = 24
        print("Default-Werte (detecto.ini):")
        print(f"  {'examplecount:':<{label_b}} {ini_defaults['examplecount']}")
        print(f"  {'minlen:':<{label_b}} {ini_defaults['minlen']}")
        print(f"  {'critical:':<{label_b}} {ini_defaults['critical']}")
        print(f"  {'anon:':<{label_b}} {ini_defaults['anon']}")
        print(f"  {'full:':<{label_b}} {ini_defaults['full']}")
        print(f"  {'nocolor:':<{label_b}} {ini_defaults['nocolor']}")
        print(f"  {'showskipped:':<{label_b}} {ini_defaults['showskipped']}")
        print(f"  {'anon_muster:':<{label_b}} {ini_defaults['anon_muster']}")
        print(f"  {'search_regexp:':<{label_b}} {ini_defaults['search_regexp']}")
        print(f"  {'search_field:':<{label_b}} {ini_defaults['search_field']}")
        print(f"  {'search_suchmuster:':<{label_b}} {ini_defaults['search_suchmuster']}")
        print()
        def krit_farbig(krit):
            farbe = FARBE_ROT if krit <= 2 else FARBE_GELB if krit == 3 else FARBE_RESET
            return f"{farbe}[{KRIT_LABELS[krit]}]{FARBE_RESET}"

        print(f"Regexp: {len(regexp_muster)}")
        for name, krit, beschreibung, _ in regexp_muster:
            print(f"  {krit_farbig(krit)} {name}: {beschreibung}")
        print(f"Field: {len(field_muster)}")
        for name, krit, beschreibung, _, offset in field_muster:
            print(f"  {krit_farbig(krit)} {name}: {beschreibung} (offset: {offset})")
        print(f"Suchmuster: {len(such_muster)}")
        for name, krit, werte in such_muster:
            print(f"  {krit_farbig(krit)} {name}: {len(werte)} Einträge")
        sys.exit(0)

    print(f"Geladene Regexp-Muster: {len(regexp_muster)}")
    print(f"Geladene Field-Muster: {len(field_muster)}")
    print(f"Geladene Suchmuster-Kategorien: {len(such_muster)}")

    # Logdateien sammeln: direkte Dateien und Glob-Muster auflösen
    logdateien = []
    for eingabe in log_eingaben:
        gefunden = finde_logdateien(eingabe)
        logdateien.extend(gefunden)
    # Duplikate entfernen, Reihenfolge beibehalten
    logdateien = list(dict.fromkeys(logdateien))
    if not logdateien:
        print(f"FEHLER: Keine Logdateien gefunden für: {' '.join(log_eingaben)}")
        sys.exit(1)

    print(f"Analysiere {len(logdateien)} Logdatei(en): {', '.join(logdateien)}")

    import time
    start_zeit = time.time()
    ergebnisse, anzahl_zeilen = analysiere_logdateien(
        logdateien, regexp_muster, field_muster, such_muster,
        sw_regexp, sw_field, sw_suchmuster)
    dauer_sek = time.time() - start_zeit

    ausgabe(ergebnisse, beispiel_anzahl, args.anon, args.full, args.nocolor,
           args.critical, args.showskipped)

    # Statistik
    minuten = int(dauer_sek // 60)
    sekunden = int(dauer_sek % 60)
    if minuten > 0:
        dauer_text = f"{minuten} min {sekunden} sec.   ({int(dauer_sek)} sec)"
    else:
        dauer_text = f"{sekunden} sec.   ({dauer_sek:.1f} sec)"
    label_breite = 24
    print("Statistik")
    print(f"{'Analysierte Logdateien:':<{label_breite}} {len(logdateien)}")
    print(f"{'Anzahl der Zeilen:':<{label_breite}} {anzahl_zeilen:,}".replace(",", "."))
    print(f"{'Analyse-Dauer:':<{label_breite}} {dauer_text}")
    print()

    # Logdatei schreiben
    if args.logresult is not None or args.logresultanon is not None:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"detecto_{timestamp}.log"
        log_anon = args.logresultanon is not None
        # Dateiname bestimmen: benutzerdefiniert oder Standard
        benutzername = (args.logresultanon if log_anon else args.logresult) or ""
        log_dateiname = benutzername if benutzername else default_name

        # Ergebniszeilen ohne Farbe erzeugen
        log_zeilen = erzeuge_ergebnis_zeilen(ergebnisse, beispiel_anzahl,
                                              anon=log_anon, full=args.full,
                                              farbe=False, critical=args.critical,
                                              showskipped=args.showskipped)

        aufruf = erzeuge_aufruf_string(args, logdateien, ini_defaults)

        with open(log_dateiname, "w", encoding="utf-8") as f:
            from datetime import date
            f.write(f"Detecto v{VERSION}  ({date.today().isoformat()})\n")
            f.write("Copyright (c) 2026 Alexander Kornbrust  |  MIT License\n")
            f.write(f"\nAufruf: {aufruf}\n")
            f.write(f"\nGeladene Regexp-Muster: {len(regexp_muster)}\n")
            f.write(f"Geladene Field-Muster: {len(field_muster)}\n")
            f.write(f"Geladene Suchmuster-Kategorien: {len(such_muster)}\n")
            f.write(f"Analysiere {len(logdateien)} Logdatei(en): "
                    f"{', '.join(logdateien)}\n")
            for zeile in log_zeilen:
                f.write(zeile + "\n")
            f.write("Statistik\n")
            f.write(f"{'Analysierte Logdateien:':<{label_breite}} "
                    f"{len(logdateien)}\n")
            f.write(f"{'Anzahl der Zeilen:':<{label_breite}} "
                    f"{anzahl_zeilen:,}\n".replace(",", "."))
            f.write(f"{'Analyse-Dauer:':<{label_breite}} {dauer_text}\n")

        print(f"Ergebnis gespeichert in: {log_dateiname}"
              f"{' (anonymisiert)' if log_anon else ''}")

    # Excel-Export
    if args.xlsx is not None:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_default = f"detecto_{timestamp}.xlsx"
        xlsx_dateiname = args.xlsx if args.xlsx else xlsx_default
        aufruf_xlsx = erzeuge_aufruf_string(args, logdateien, ini_defaults)
        exportiere_xlsx(
            xlsx_dateiname, ergebnisse, beispiel_anzahl, args.critical,
            aufruf_xlsx, logdateien, anzahl_zeilen, dauer_text,
            regexp_muster, field_muster, such_muster, args.full,
            args.excelanon)


if __name__ == "__main__":
    main()
