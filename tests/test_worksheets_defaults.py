"""Findings 36, 38, 39: worksheet count and default consistency."""
from __future__ import annotations

import configparser
import re
from collections import OrderedDict
from importlib.resources import files
from pathlib import Path

import pytest

from detecto.anonymizer import Anonymizer
from detecto.config import DetectoConfig
from detecto.exporter import ExportContext, export_xlsx


def _ctx(full):
    results: OrderedDict = OrderedDict()
    results["email"] = ("regexp", 4, {"a@b.example": [("t.log", "a@b.example", None)]})
    return ExportContext(
        results=results, example_count=3, critical=5, call_str="t",
        logfiles=["t.log"], line_count=1, duration_text="0 sec",
        regexp=[("email", 4, "E", re.compile(r"\S+@\S+"), "token")],
        field=[], search=[], full=full, anonymizer=Anonymizer(),
    )


# --- Finding 36: worksheet count ------------------------------------------

def test_five_standard_worksheets(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    out = tmp_path / "n.xlsx"
    export_xlsx(str(out), _ctx(full=False))
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Findings", "Tool", "Regexp", "Field", "String"]
    assert len(wb.sheetnames) == 5


def test_optional_full_worksheet(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    out = tmp_path / "f.xlsx"
    export_xlsx(str(out), _ctx(full=True))
    wb = openpyxl.load_workbook(out)
    assert "Full" in wb.sheetnames
    assert len(wb.sheetnames) == 6


# --- Findings 38/39: code defaults == shipped INI -------------------------

def _shipped_ini_defaults():
    data = Path(str(files("detecto") / "data"))
    cp = configparser.ConfigParser()
    cp.read(str(data / "detecto.ini"), encoding="utf-8")
    return cp["defaults"]


def test_examplecount_default_consistent():
    ini = _shipped_ini_defaults()
    assert DetectoConfig().examplecount == int(ini["examplecount"])


def test_minlen_default_consistent():
    ini = _shipped_ini_defaults()
    assert DetectoConfig().minlen == int(ini["minlen"])


# --- Finding 37: IBAN criticality unified ---------------------------------

def test_iban_criticality_is_two():
    data = Path(str(files("detecto") / "data"))
    for line in (data / "regexp.csv").read_text(encoding="utf-8").splitlines():
        if line.startswith("IBAN::"):
            assert line.split("::")[1] == "2"
            break
    else:
        pytest.fail("IBAN pattern not found")
