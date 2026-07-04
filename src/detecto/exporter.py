"""Excel and log file export for Detecto results."""
from __future__ import annotations

import logging
import os
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, datetime

from detecto import VERSION
from detecto.anonymizer import Anonymizer
from detecto.constants import (
    COPYRIGHT_YEAR, LABEL_WIDTH, EXCEL_FORMULA_CHARS,
    CLR_HEADER, CLR_CUSTOMER, CLR_VALUE, CLR_FIELD, CLR_KRIT,
)
from detecto.diagnostics import ScanDiagnostics
from detecto.formatter import build_result_lines
from detecto.loaders import RegexpPattern, FieldPattern, SearchPattern

__all__ = ["export_log", "export_xlsx", "ExportContext"]

log = logging.getLogger(__name__)


@dataclass
class ExportContext:
    """Bundles all parameters needed for log/Excel export."""

    results: OrderedDict
    example_count: int
    critical: int
    call_str: str
    logfiles: list[str]
    line_count: int
    duration_text: str
    regexp: list[RegexpPattern]
    field: list[FieldPattern]
    search: list[SearchPattern]
    full: bool = False
    excelanon: bool = False
    anonymizer: Anonymizer | None = None
    diagnostics: ScanDiagnostics | None = None


def _restrict_permissions(path: str) -> None:
    """Restrict a freshly written report to the current user (Finding 3).

    Report files can contain sensitive findings, so on POSIX we set 0600.
    Failure (e.g. on Windows or an unusual filesystem) is only a warning.
    """
    if os.name != "posix":
        return
    try:
        os.chmod(path, 0o600)
    except OSError as e:  # pragma: no cover - platform dependent
        log.warning("Could not restrict permissions on %s: %s", path, e)


def _sanitize_cell(value: object) -> object:
    """Prevent Excel formula injection by prefixing dangerous strings."""
    if isinstance(value, str) and value and value[0] in EXCEL_FORMULA_CHARS:
        return f"'{value}"
    return value


class _SheetBuilder:
    """DRY helper for creating formatted Excel worksheets."""

    def __init__(self, wb: object) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        self.wb = wb
        self._h_font = Font(bold=True, color="FFFFFF", size=11)
        self._h_fill = PatternFill(
            start_color=CLR_HEADER, end_color=CLR_HEADER, fill_type="solid"
        )
        self._h_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        self._bold = Font(bold=True)
        self._center = Alignment(horizontal="center")
        self._wrap_top = Alignment(wrap_text=True, vertical="top")
        self.krit_fills = {
            k: PatternFill(start_color=c, end_color=c, fill_type="solid")
            for k, c in CLR_KRIT.items()
        }

    def make_fill(self, color: str) -> object:
        """Create a PatternFill from an ARGB color string."""
        from openpyxl.styles import PatternFill
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def create_sheet(
        self, title: str, headers: list[str], widths: list[int],
        *, index: int | None = None, use_active: bool = False,
    ) -> object:
        """Create a sheet with formatted header, column widths and frozen top row."""
        if use_active:
            ws = self.wb.active
            ws.title = title
        elif index is not None:
            ws = self.wb.create_sheet(title=title, index=index)
        else:
            ws = self.wb.create_sheet(title=title)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = self._h_font
            c.fill = self._h_fill
            c.alignment = self._h_align
            c.border = self.border
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
        ws.freeze_panes = "A2"
        return ws

    def add_rows(
        self, ws: object, data: list[list[object]], last_col: str,
        *, krit_col: int | None = None, center_cols: list[int] | None = None,
        bold_cols: list[int] | None = None, wrap_cols: list[int] | None = None,
        fill_cols: dict[int, object] | None = None,
    ) -> int:
        """Write data rows with formatting and set autofilter. Returns last row."""
        center = set(center_cols or [])
        bold = set(bold_cols or [])
        wrap = set(wrap_cols or [])
        fills = fill_cols or {}
        row = 2
        for values in data:
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=_sanitize_cell(val))
                cell.border = self.border
                if krit_col and col == krit_col:
                    cell.fill = self.krit_fills.get(val, self.krit_fills[5])
                    cell.alignment = self._center
                if col in center:
                    cell.alignment = self._center
                if col in bold:
                    cell.font = self._bold
                if col in wrap:
                    cell.alignment = self._wrap_top
                if col in fills:
                    cell.fill = fills[col]
            row += 1
        ws.auto_filter.ref = f"A1:{last_col}{max(row - 1, 1)}"
        return row - 1


def export_log(
    filename: str, ctx: ExportContext,
    anon: bool = False, show_skipped: bool = False,
) -> None:
    """Save Detecto results as a plain-text log file."""
    lines = build_result_lines(
        ctx.results, ctx.example_count, anon=anon, full=ctx.full,
        color=False, critical=ctx.critical, show_skipped=show_skipped,
        anonymizer=ctx.anonymizer if anon else None,
    )
    with open(filename, "w", encoding="utf-8") as f:
        f.write(f"Detecto v{VERSION}  ({date.today().isoformat()})\n")
        f.write(f"Copyright (c) {COPYRIGHT_YEAR} Alexander Kornbrust  |  MIT License\n")
        f.write(f"\nAufruf: detecto {ctx.call_str}\n")
        f.write(f"\nGeladene Regexp-Muster: {len(ctx.regexp)}\n")
        f.write(f"Geladene Field-Muster: {len(ctx.field)}\n")
        f.write(f"Geladene Suchmuster-Kategorien: {len(ctx.search)}\n")
        f.write(f"Analysiere {len(ctx.logfiles)} Logdatei(en): "
                f"{', '.join(ctx.logfiles)}\n")
        for line in lines:
            f.write(line + "\n")
        f.write("Statistik\n")
        f.write(f"{'Analysierte Logdateien:':<{LABEL_WIDTH}} {len(ctx.logfiles)}\n")
        f.write(f"{'Anzahl der Zeilen:':<{LABEL_WIDTH}} "
                f"{ctx.line_count:,}\n".replace(",", "."))
        f.write(f"{'Analyse-Dauer:':<{LABEL_WIDTH}} {ctx.duration_text}\n")
        if ctx.diagnostics is not None:
            f.write("\n")
            for sline in ctx.diagnostics.summary_lines():
                f.write(sline + "\n")
    _restrict_permissions(filename)
    log.info("Log result saved: %s", filename)


def export_xlsx(filename: str, ctx: ExportContext) -> None:
    """Export findings as an Excel file (plain text, no Rich Text)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        log.error("openpyxl not installed. pip install openpyxl")
        return

    wb = Workbook()
    sb = _SheetBuilder(wb)
    art_order = {"field": 0, "regexp": 1, "string": 2}
    sorted_items = sorted(
        ctx.results.items(),
        key=lambda x: (x[1][1], art_order.get(x[1][0], 9), x[0]),
    )

    _build_findings(sb, sorted_items, ctx)
    if ctx.full:
        _build_full(sb, sorted_items, ctx)
    _build_tool(sb, ctx)
    _build_patterns(sb, ctx.regexp, ctx.field, ctx.search)

    try:
        wb.save(filename)
    except (IOError, OSError) as e:
        log.error("Failed to save Excel file %s: %s", filename, e)
        print(f"FEHLER: Excel-Datei konnte nicht gespeichert werden: {e}",
              file=sys.stderr)
        return
    _restrict_permissions(filename)
    hint = " (anonymisiert)" if ctx.excelanon else ""
    print(f"Excel-Report gespeichert in: {filename}{hint}")


def _build_findings(sb: _SheetBuilder, sorted_items: list, ctx: ExportContext) -> None:
    ws = sb.create_sheet(
        "Findings",
        ["Bearbeiter", "Finding", "Kommentar", "Art", "Muster",
         "Kritikalit\u00e4t (1-5)", "Beispiele", "gefunden in"],
        [15, 15, 25, 10, 20, 18, 60, 53], use_active=True,
    )
    cust_fill = sb.make_fill(CLR_CUSTOMER)
    rows: list[list] = []
    for name, (typ, krit, hits) in sorted_items:
        if not hits or krit > ctx.critical:
            continue
        keys = sorted(hits.keys())[: ctx.example_count]
        examples = (
            ", ".join(ctx.anonymizer.redact(k) for k in keys)
            if ctx.excelanon and ctx.anonymizer
            else ", ".join(keys)
        )
        logs = sorted({e[0] for vals in hits.values() for e in vals})
        rows.append(["", "", "", typ, name, krit, examples, ", ".join(logs)])
    sb.add_rows(
        ws, rows, "H", krit_col=6, wrap_cols=[7],
        fill_cols={1: cust_fill, 2: cust_fill, 3: cust_fill},
    )


def _build_full(sb: _SheetBuilder, sorted_items: list, ctx: ExportContext) -> None:
    from openpyxl.styles import Font

    ws = sb.create_sheet(
        "Full",
        ["Logdatei", "Art", "Muster", "Kritikalit\u00e4t (1-5)",
         "Gefundener Wert", "Feld", "Logeintrag"],
        [25, 10, 20, 18, 30, 25, 100], index=1,
    )
    val_fill = sb.make_fill(CLR_VALUE)
    fld_fill = sb.make_fill(CLR_FIELD)
    rows: list[list] = []
    field_rows: list[int] = []
    for name, (typ, krit, hits) in sorted_items:
        if not hits or krit > ctx.critical:
            continue
        for value in sorted(hits.keys())[: ctx.example_count]:
            entry = hits[value][0]
            ftk = entry[2] if len(entry) > 2 else None
            # Finding 7/24: prefer the original value + span over the key.
            orig = entry[3] if len(entry) > 3 else value
            start = entry[4] if len(entry) > 4 else -1
            end = entry[5] if len(entry) > 5 else -1
            display_val = (
                ctx.anonymizer.redact(orig)
                if ctx.excelanon and ctx.anonymizer
                else orig
            )
            marked = _mark_log(entry[1], orig, ftk, ctx.excelanon,
                               ctx.anonymizer, start, end)
            rows.append([entry[0], typ, name, krit, display_val, ftk or "", marked])
            if ftk:
                field_rows.append(len(rows) - 1)
    sb.add_rows(ws, rows, "G", krit_col=4, bold_cols=[5], wrap_cols=[7],
                fill_cols={5: val_fill})
    for idx in field_rows:
        fc = ws.cell(row=idx + 2, column=6)
        fc.fill = fld_fill
        fc.font = Font(bold=True)


def _mark_log(
    log_line: str, value: str, field_token: str | None,
    excelanon: bool, anon: Anonymizer | None,
    start: int = -1, end: int = -1,
) -> str:
    """Mark matches (>>>value<<<) and fields ([FELD:name]) in a log entry.

    Finding 24: ``value`` is the exact original text; when a reliable span is
    available and no redaction is needed, the exact position is used.
    """
    if not value:
        return log_line
    ml = log_line
    if field_token:
        ml = re.sub(
            re.escape(field_token),
            lambda m: f"[FELD:{m.group(0)}]", ml,
            count=1, flags=re.IGNORECASE,
        )
    replacement = anon.redact(value) if excelanon and anon else value
    # Replacements via lambda: values with backslashes (e.g. 'DOMAIN\\user')
    # would otherwise raise re.error and silently drop the markers.
    return re.sub(
        re.escape(value),
        lambda m: f">>>{replacement}<<<", ml,
        flags=re.IGNORECASE,
    )


def _build_tool(sb: _SheetBuilder, ctx: ExportContext) -> None:
    from openpyxl.styles import Font

    ws = sb.wb.create_sheet(title="Tool")
    bold = Font(bold=True, size=11)
    data = [
        ("Detecto Version", f"v{VERSION}"),
        ("Datum", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Aufruf", f"detecto {ctx.call_str}"),
        ("", ""),
        ("Analysierte Logdateien", str(len(ctx.logfiles))),
        ("Logdateien", ", ".join(ctx.logfiles)),
        ("Anzahl der Zeilen", str(ctx.line_count)),
        ("Analyse-Dauer", ctx.duration_text),
        ("", ""),
        ("Geladene Regexp-Muster", str(len(ctx.regexp))),
        ("Geladene Field-Muster", str(len(ctx.field))),
        ("Geladene Suchmuster-Kategorien", str(len(ctx.search))),
        ("Critical-Filter", str(ctx.critical)),
        ("Beispiele pro Typ", str(ctx.example_count)),
    ]
    if ctx.diagnostics is not None:
        d = ctx.diagnostics
        data.extend([
            ("", ""),
            ("Scan-Status", d.status().value),
            ("Dateien vollstaendig", str(d.files_complete)),
            ("Dateien teilweise", str(d.files_partial)),
            ("Dateien nicht lesbar", str(d.files_unreadable)),
            ("Sonstige Scanfehler", str(d.other_errors)),
            ("Regex-Timeouts", str(sum(d.regex_timeouts.values()))),
            ("Deaktivierte Muster", ", ".join(sorted(d.disabled_patterns)) or "-"),
            ("Verworfen (Limits)",
             str(d.findings_dropped_global + d.findings_dropped_pattern)),
        ])
    for r, (label, val) in enumerate(data, 1):
        ws.cell(row=r, column=1, value=_sanitize_cell(label)).font = bold
        ws.cell(row=r, column=2, value=_sanitize_cell(val))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def _build_patterns(
    sb: _SheetBuilder,
    regexp: list[RegexpPattern],
    field: list[FieldPattern],
    search: list[SearchPattern],
) -> None:
    ws = sb.create_sheet(
        "Regexp",
        ["Name", "Kritikalit\u00e4t", "Beschreibung", "Pattern"],
        [20, 14, 40, 60],
    )
    sb.add_rows(ws, [[n, k, d, p.pattern] for n, k, d, p in regexp], "D", krit_col=2)

    ws = sb.create_sheet(
        "Field",
        ["Name", "Kritikalit\u00e4t", "Beschreibung", "Pattern", "Offset"],
        [20, 14, 35, 60, 10],
    )
    sb.add_rows(
        ws, [[n, k, d, p.pattern, o] for n, k, d, p, o in field],
        "E", krit_col=2, center_cols=[5],
    )

    ws = sb.create_sheet(
        "String",
        ["Name", "Kritikalit\u00e4t", "Eintr\u00e4ge"],
        [20, 14, 12],
    )
    # Kein geratener Dateiname mehr (Kategorie 'Ort' lag in 'orte.csv',
    # nicht 'ort.csv') - die Zuordnung steht in suchmuster.csv.
    sb.add_rows(
        ws, [[n, k, len(v)] for n, k, v in search],
        "C", krit_col=2, center_cols=[3],
    )
